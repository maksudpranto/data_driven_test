import openpyxl


def row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


def column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


def read_data(file, sheet, reading_rows, reading_column):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_rows, column=reading_column).value


def write_data(file, sheet, writing_row, writing_column, data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    worksheet.cell(row=writing_row, column=writing_column).value = data
    workbook.save(file)
